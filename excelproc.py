from operatortype import OperatorType
from tkinter import filedialog
from commontools import CommonTools
import constant

import os
from openpyxl import Workbook
from openpyxl.styles import Alignment

class ExcelProc():
    def __init__(self, operatorType: OperatorType, folder: str=None):
        """
        导出 Excel 文件

        Params:
            operatorType (OperatorType): 操作类型
            folder (str): 数据来源文件夹路径
        """
        self.operatorType = operatorType
        self.folder = folder
        self.filePath = None
        self.wb = None
        self.ws = None
        self.row = 1

    def openExcel(self) -> str:
        """
        打开 Excel 文件

        Returns:
            str: 保存的文件路径；None 未选择保存路径
        """
        if self.operatorType == OperatorType.FOLDER:
            initialfile = CommonTools.getNameFromPath(self.folder) + ' ' + CommonTools.getTime() + '.xlsx'
        else:
            self.folder = constant.Path.programFileDir
            initialfile = CommonTools.getNameFromPath(self.folder) + ' ' + CommonTools.getTime() + '.xlsx'

        self.filePath = filedialog.asksaveasfilename(
            initialdir=self.folder,
            initialfile=initialfile,
            title='保存文件',
            filetypes=[('Excel 工作簿', '*.xlsx')]
        )
        if self.filePath:
            self.filePath = os.path.normpath(self.filePath)
            self.wb = Workbook()
            self.ws = self.wb.active
            return self.filePath
        else:
            return None
        
    def writeContent(self, content: dict, fromPath: str) -> tuple[bool, str]:
        """
        写入内容

        Params:
            content (dict): 内容
            fromPath (str): 内容来源文件路径

        Returns:
            tuple[bool, str]: (是否写入成功, 失败原因)
        """
        if self.ws is None:
            return False, '未打开 Excel 文件'
        if self.row == 1:
            self.ws.cell(row=self.row, column=1, value='序号')
            self.ws.cell(row=self.row, column=2, value='通报名称')
            self.ws.cell(row=self.row, column=3, value='发生日期')
            self.ws.cell(row=self.row, column=4, value='问题描述')
            self.ws.cell(row=self.row, column=5, value='责任人')
            self.ws.cell(row=self.row, column=6, value='通报日期')
            self.ws.cell(row=self.row, column=7, value='责任追究处理')
            self.row += 1

        self.ws.column_dimensions['A'].width = 20  # 序号
        self.ws.column_dimensions['B'].width = 45  # 通报名称
        self.ws.column_dimensions['C'].width = 10  # 发生日期
        self.ws.column_dimensions['D'].width = 70  # 问题描述
        self.ws.column_dimensions['E'].width = 10  # 责任人
        self.ws.column_dimensions['F'].width = 10  # 通报日期
        self.ws.column_dimensions['G'].width = 50  # 责任追究处理
        
        serialNumber = content['序号']
        reportName = content['通报名称']
        occurDate = content['发生日期']
        problemDescription = content['问题描述']
        responsiblePerson = content['责任人']
        reportDate = content['通报日期']
        responsibilityInvestigation = content['责任追究处理']

        def writeLine():
            self.ws.cell(row=self.row, column=1, value=serialNumber).alignment = Alignment(wrap_text=True, vertical='center')
            file = self.ws.cell(row=self.row, column=2, value=reportName)
            file.hyperlink = fromPath
            file.style = 'Hyperlink'
            file.alignment = Alignment(wrap_text=True, vertical='center')
            self.ws.cell(row=self.row, column=3, value=occurDate).alignment = Alignment(wrap_text=True, vertical='center')
            self.ws.cell(row=self.row, column=4, value=problemDescription).alignment = Alignment(wrap_text=True, vertical='center')
            self.ws.cell(row=self.row, column=6, value=reportDate).alignment = Alignment(wrap_text=True, vertical='center')
            self.ws.cell(row=self.row, column=7, value=responsibilityInvestigation).alignment = Alignment(wrap_text=True, vertical='center')

        writeLine()
        if len(responsiblePerson) != 0:
            self.ws.cell(row=self.row, column=5, value=responsiblePerson[0]).alignment = Alignment(wrap_text=True, vertical='center')
        self.row += 1        

        if len(responsiblePerson) >= 2:
            for count in range(1, len(responsiblePerson)):
                writeLine()
                self.ws.cell(row=self.row, column=5, value=responsiblePerson[count]).alignment = Alignment(wrap_text=True, vertical='center')
                self.row += 1
    def closeExcel(self):
        """
        关闭并保存文件
        """
        self.wb.save(self.filePath)
        self.wb = None
        self.ws = None
        self.row = 1
        os.startfile(self.filePath)